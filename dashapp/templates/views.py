from django.shortcuts import render
from django.db.models import Count
from dashapp.models import Dashboard_Info,ProjectDetails,RagDetails,ProjectReview,Resources,resourceSkills
from dashapp.forms import Dashboard_InfoForm,ProjectReviewForm
from dashapp.models import EntTechnologies,SpTechnologies,ScriptingSkills,Automation1
from django.http import HttpResponse
from django.shortcuts import redirect
import datetime
import openpyxl
from dashapp.db import insertDB, connectDB,selectDB,getIndex,listDB,deleteDB,updateDB,selectDBParam
from django.contrib import messages
import json
from collections import Counter
from operator import add
from dashapp.generalUtil import returnStatusValue
from datetime import datetime
# Create your views here.

def addDash(request):
    return render(request,'addProj.html')

def addProject(request):
    if request.method == "POST":
        form = Dashboard_InfoForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('/mainpage')
            except:
                return HttpResponse("<h3> OOPS! Something went wrong. Please try again <\h3>")
    else:
        form = Dashboard_InfoForm()
    return render(request, '',{'form':form})

def dashPageShow(request):

    dbi  = Dashboard_Info.objects.all()
    entRows = Dashboard_Info.objects.filter(category = 'ENT')
    spRows = Dashboard_Info.objects.filter(category = 'SP')
    greenRows = Dashboard_Info.objects.filter(completionZone = 'Green')
    amberRows = Dashboard_Info.objects.filter(completionZone = 'Amber')
    redRows = Dashboard_Info.objects.filter(completionZone = 'Red')
    newProjRows = Dashboard_Info.objects.filter(projectStatus = 'New')
    currentProjRows = Dashboard_Info.objects.filter(projectStatus = 'On Going')
    completedProjRows = Dashboard_Info.objects.filter(projectStatus = 'Completed')
    entCount = entRows.count()
    spCount = spRows.count()
    greenCount = greenRows.count()
    amberCount = amberRows.count()
    redCount = redRows.count()
    newCount = newProjRows.count()
    currentCount = currentProjRows.count()
    completedCount = completedProjRows.count()
    projDetails = ProjectDetails.objects.all()
    print(dbi[0])
    #print(projDetails)
    #print(projDetails[0].dealValue)
    #print(projDetails[0].projectName)
    #print(greenCount)
    return render(request,'mainPage.html',{'dbi':dbi, 'entCount':entCount, 'spCount':spCount, 'greenCount':greenCount,
    'amberCount':amberCount,'redCount':redCount,'newCount':newCount,'currentCount':currentCount,'completedCount':completedCount,
    'projDetails':projDetails})

def editDashboardInfo(request, projectName):
    dash = Dashboard_Info.objects.get( projectName = projectName)
    return render(request, 'editDashPage.html',{'dash':dash})

def updateDashboardInfo(request, projectName):
    dash  = Dashboard_Info.objects.get(projectName = projectName)
    form = Dashboard_InfoForm(request.POST, instance = dash)
    print(form)
    if form.is_valid():
        form.save()
        return redirect('/mainpage')
    #return render(request,"dashPageInfo.html",{'dash':dash})

def deleteDashboardInfo(request, projectName):
    dash = Dashboard_Info.objects.get(projectName = projectName)
    dash.delete()
    return redirect('/mainpage')

def showprojectDetails(request, projectName):
    projDetInfo = ProjectDetails.objects.filter(projectName = projectName)
    tSites = projDetInfo[0].totalSites
    projRagInfo = RagDetails.objects.filter(ProjectId = projDetInfo[0].ProjectId)
    winCompletedList = list()
    monthList = list()
    siteComp = 0

    for ragRows in projRagInfo:
        #year = datetime.datetime.strptime(str(ragRows),'%Y-%m')
        siteComp = siteComp + ragRows.sitesCompleted
        month = str(ragRows.month)
        year = str(ragRows.year)
        mli = year+'-'+month
        winCompletedList.append(ragRows.windowsCompleted)
        monthList.append(ragRows.month)

    if projDetInfo.count() != 0:
        completedPercent = round((siteComp/tSites)*100)
        sitesInProgress = tSites - siteComp
        sitesInProgress = round((sitesInProgress/tSites)*100)
        projDetails = ProjectDetails.objects.get(ProjectId = projDetInfo[0].ProjectId)
        ragDetails = RagDetails.objects.filter(ProjectId = projDetInfo[0].ProjectId)
        completedSites = list()
        xAxis = list()
        custSatisfaction = list()
        windowsPlanned = list()
        windowsCompleted = list()
        
        for rag in ragDetails:
            completedSites.append(rag.sitesCompleted)
            xAxis.append(str(rag.year)+' Month '+str(rag.month))
            custSatisfaction.append(rag.custSatisfaction)
            windowsPlanned.append(rag.windowsPlanned)
            windowsCompleted.append(rag.windowsCompleted)
            
       
        #print(completedSites)
        #print(xAxis)
        completedSites = json.dumps(completedSites)
        custSatisfaction = json.dumps(custSatisfaction)
        xAxis = json.dumps(xAxis)

        #print(windowsPlanned)
        #print(windowsCompleted)

        windowSuccessRate = list()
        
        
        for i in range(len(windowsPlanned)):
            #print(float(windowsCompleted[i]))
            temp1 = float(windowsCompleted[i])
            #print(temp1)
            print(windowsPlanned[i])
            print(windowsCompleted[i])
            if windowsCompleted[i] > windowsPlanned[i]:
                windowSuccessRate.append(100)
            else:
                windowSuccessRate.append((temp1/windowsPlanned[i])*100)

            #print(temp1/windowsPlanned[i])
            #print((float64(windowsCompleted[i])//float64(windowsPlanned[i])))
            #print(windowsCompleted[i])

        

        windowsPlanned = json.dumps(windowsPlanned)
        windowsCompleted = json.dumps(windowsCompleted)
        windowSuccessRate = json.dumps(windowSuccessRate)
        
        #print(windowSuccessRate)
        #print(completedSites)
        #print(xAxis)
        #print(projDetInfo[0].ProjectId)
        
        

        #for i in range(len(windowsPlanned)):
         #   print(windowsPlanned[i])
          #  print(windowsCompleted[i])

        return render(request, 'projectDetails.html', {'pdInfo': projDetInfo[0], 'winCompletedList':winCompletedList,'monthList':monthList,
        'siteComp':siteComp, 'tSites':tSites,'projectName':projectName,
        'completedPercent':completedPercent,'sitesInProgress':sitesInProgress,
        'projDetails':projDetails,'ragDetails':ragDetails,'xAxis':xAxis,'completedSites':completedSites,
        'custSatisfaction':custSatisfaction,'windowsPlanned':windowsPlanned,
        'windowsCompleted':windowsCompleted,'windowSuccessRate':windowSuccessRate})
        #return render(request, 'rough.html', {'pdInfo': projDetInfo[0], 'winCompletedList':winCompletedList,'monthList':monthList,
        #'siteComp':siteComp, 'tSites':tSites})
    else:
        return HttpResponse("<h3> OOPS! The requested Project '"+projectName+"' does not have any data yet")

def projAdditionalDetails(request, projectId):
    projDetails = ProjectDetails.objects.get(ProjectId = projectId)
    ragDetails = RagDetails.objects.filter(ProjectId = projectId)

    return render(request,'projectAdditionalInfo.html',{'pdInfo':projDetails,'ragDetails':ragDetails})

def uploadProjectDetails(request):
    if "GET" == request.method:
        return render(request, 'uploadProjectDetails.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        worksheet2 = wb["Sheet2"]
        #print(worksheet)

        excel_data = list()
        excel_data_sheet2 = list()
        # iterating over the rows and
        # getting value from each cell in row



        for i,row in enumerate(worksheet.iter_rows()):
            row_data = list()
            if i == 0:
                continue
            for cell in row:
                row_data.append((cell.value))
            excel_data.append(row_data)
            '''(cursor, cnx) = connectDB()
            cursor.execute("SELECT * FROM projectdetails;")
            rows = cursor.fetchall()
            lastid = cursor.lastrowid
            print(len(rows))
            cnx.close()'''
            #cursor.close()
            '''
            (ProjectId,category,startDate,endDate,tenure,numberOfSitesPlanned,numberOfDevicesMigrated,totalHoursAllocated,totalProjectEfforts,plannedResourceAllocation,summary,completionStatus,projectName_id,totalSites,competativeMigration,dealValue,migrationType,numOfDevicesPlannedForMigration,numOfSitesMigrated,numOfWindowsExecutedWidOutRollback,numberOfAgreedServices,numberOfServicesUpPostMigration,numberOfWindowsExecuted,numberOfWindowsPlanned,preSales,projectDeliveryManager,projectLead,projectManager,region,repeatBusiness,segment,theater,thirdPartyToCisco)

            '''
            query = ("INSERT INTO projectdetails values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);")
            val = insertDB(query,row_data)

        ragQuery = ("INSERT INTO ragdetails values (%s,%s,%s,%s,%s,%s,%s,%s,%s);")
        #for sheet 2
        for i,row in enumerate(worksheet2.iter_rows()):
            row_data2 = list()
            if i == 0:
                continue
            rowIndex = getIndex()
            row_data2.append(rowIndex+1)
            for cell in row:
                row_data2.append((cell.value))
            #excel_data.append(row_data)
            val = insertDB(ragQuery,row_data2)
            print(val)
            if val:
                messages.success(request, "Successfully Uploaded")
                #return render(request, 'uploadProjectDetails.html', {"excel_data":excel_data})
            else:
                messages.error(request, "Monthly Data ProjectId already exists!, please try again")
                #return render(request, 'uploadProjectDetails.html', {"excel_data":excel_data})
            #print("*******************")
            #print(val)

        return render(request, 'uploadProjectDetails.html', {"excel_data":excel_data})

def addEditProjReview(request, projectName):
    pri = ProjectReview.objects.filter(projectName = projectName)
    print(len(pri))
    pri_size = len(pri)
    if pri_size != 0:
        cust_rating = pri[pri_size-1].customerSentiment
        overall_rating = pri[pri_size-1].overallStatus

        completionZone = ""

        if cust_rating == "Green" and overall_rating == "Green":
            completionZone = "Green"
        elif cust_rating == "Green" and overall_rating == "Amber" or cust_rating == "Amber" and overall_rating == "Green" or cust_rating == "Amber" and overall_rating == "Amber":
            completionZone = "Amber"
        else:
            completionZone = "Red"

        Dashboard_Info.objects.filter(projectName = projectName).update(completionZone = completionZone)

        pri_reverse = reversed(pri)

        return render(request, 'showProjectReview.html', {"proj_review":pri_reverse,"proj_name":pri[0].projectName_id })
    else:
        return render(request, 'addFirstProjectReview.html',{"projectName":projectName})

def navigatetoReviewform(request, projectName):
    return render(request,'addProjectReview.html',{'projectName':projectName})

def ResourcesInfo(request):
    totalBuleBadges = Resources.objects.filter(badge = 'Blue')
    totalRedBadges = Resources.objects.filter(badge = 'Red')
    blueCount = totalBuleBadges.count()
    redCount = totalRedBadges.count()
    count = [redCount,blueCount]
    projectsInfo = Resources.objects.order_by().values('assignedProject').distinct()
    #skills = resourceSkills.objects.order_by().values('technology')

    certiDevnetCount = resourceSkills.objects.filter(certiDevnet = 'Yes').count()
    certiCcnaCount = resourceSkills.objects.filter(certiCcna = 'Yes').count()
    certiCcnpCount = resourceSkills.objects.filter(certiCcnp = 'Yes').count()
    certiSdwanCount = resourceSkills.objects.filter(certiSdwan = 'Yes').count()
    certiSdaCount = resourceSkills.objects.filter(certiSda = 'Yes').count()
    certiCcieCount = resourceSkills.objects.filter(certiCcie = 'Yes').count()
    
    certiCount = [certiDevnetCount,certiCcnaCount,certiCcnpCount,certiSdwanCount,certiSdaCount,certiCcieCount]
    certificates = ['DevNet Asc','CCNA','CCNP','SDWAN','SDA','CCIE']
    entSp = ['Enterprise Skiled','Service Provider Skilled', 'Automation Skilled']
    entSp = json.dumps(entSp)
    certiCount = json.dumps(certiCount)
    certificates = json.dumps(certificates)

    skillsList = list()
    
    entSkilledCount = EntTechnologies.objects.all().count()
    spSkilledCount = SpTechnologies.objects.all().count()
    scriptsSkilledCount = ScriptingSkills.objects.all().count()
    spEntCounts = [entSkilledCount,spSkilledCount, scriptsSkilledCount]
    spEntCounts = json.dumps(spEntCounts)
    #for skill in skills:
    #    skillsList.append(skill['technology'])
    #skillDict = dict(Counter(skillsList))
    
    #technologies = json.dumps(list(skillDict.keys()))
    #technoCount = json.dumps(list(skillDict.values()))
    
    projects = list()
    resourceCount = list()
    notAssigned = 0
    totalAssigned = 0
    for project in projectsInfo:

        if project['assignedProject'] == 'Not Assigned':
            notAssigned = Resources.objects.filter(assignedProject = project['assignedProject']).count()
            #projects.append(project['assignedProject'])
            #resourceCount.append(Resources.objects.filter(assignedProject = project['assignedProject']).count())
        else:
            totalAssigned += Resources.objects.filter(assignedProject = project['assignedProject']).count()
        projects.append(project['assignedProject'])
        resourceCount.append(Resources.objects.filter(assignedProject = project['assignedProject']).count())

    
    nested = [[]]*(len(project)-1)

    for i in range(len(projects)):
        lis = [projects[i],resourceCount[i]]
        nested.append(lis)


    assignedNotAssignedLabel = ['Total Resources', 'Assigned To Projects', 'Not Assigned To Projects']
    assignedNotAssignedCount = [totalAssigned+notAssigned, totalAssigned ,notAssigned]
    
    #nested = json.dumps(nested)
    count = json.dumps(count)
    projects = json.dumps(projects)
    resourceCount = json.dumps(resourceCount)
    assignedNotAssignedLabel = json.dumps(assignedNotAssignedLabel)
    assignedNotAssignedCount = json.dumps(assignedNotAssignedCount)
    #technologies':technologies
    return render(request, 'resources.html',{'blueCount':blueCount,'redCount':redCount,
        'projects':projects,'resourceCount':resourceCount,'nestedData':nested,'count':count,
        'certiCount':certiCount,'certificates':certificates,'entSp':entSp,'spEntCounts':spEntCounts,
        'assignedNotAssignedLabel':assignedNotAssignedLabel,'assignedNotAssignedCount':assignedNotAssignedCount})

def insertReview(request, projectName):
    if request.method == "POST":
        form = ProjectReviewForm(request.POST)
        print(form.is_valid())
        if form.is_valid():
            print(form)
            try:
                index_val = form.save()
                #print(index_val)
                schedule_status = form.cleaned_data.get('scheduleStatus')
                quality_status = form.cleaned_data.get('qualityStatus')
                resource_status = form.cleaned_data.get('resourceStatus')
                automation_status = form.cleaned_data.get('automationStatus')

                overall_status = ""
                green = "Green"
                amber = "Amber"
                red = "Red"

                if schedule_status == green and quality_status == green and resource_status == green and automation_status == green:
                    overall_status = green
                elif schedule_status == amber and quality_status == green and resource_status == green and automation_status == green or schedule_status == green and quality_status == amber and resource_status == green and automation_status == green or schedule_status == green and quality_status == green and resource_status == amber and automation_status == green or schedule_status == green and quality_status == green and resource_status == green and automation_status == amber:
                    overall_status = amber
                elif schedule_status == red or quality_status == red or resource_status == red or automation_status == red:
                    overall_status = red
                    
                ProjectReview.objects.filter(id = index_val.id).update(overallStatus = overall_status)

                return redirect('/showProjectReview/'+projectName)
            except:
                return HttpResponse("<h3> OOPS! Something went wrong. Please try again <\h3>")

    else:
        form = ProjectReviewForm()
    return render(request, '',{'form':form})

def redBlueResources(request):
    totalBlueCount = Resources.objects.filter(badge='Blue').count()
    blueProjects = Resources.objects.filter(badge='Blue')
    tempProj = blueProjects.order_by().values('assignedProject').distinct()
    entCount = blueProjects.filter(entOrSp='ENT').count()
    spCount = blueProjects.filter(entOrSp='SP').count()
    notAlignedToProjects = blueProjects.filter(assignedProject = 'Not Assigned')
    notAlignedToProjectsCount = notAlignedToProjects.count()
    blueLeads = blueProjects.filter(projectRole = 'Lead')
    leadsCapable = blueProjects.filter(leadCapable = 'Yes')
    
    blueLeadsCount = blueLeads.count()
    leadsCapableCount = leadsCapable.count()

    #projects = Resources.objects.order_by().values('assignedProject').distinct()
    projects = list()
    for pr in tempProj:
        if pr['assignedProject'] == 'Not Assigned':
            continue
        projects.append(pr['assignedProject'])
    
    notInProjName = list()
    notInProjId = list()
    
    for notAligned in notAlignedToProjects:
        notInProjName.append(notAligned.name)
        notInProjId.append(notAligned.cecID)

    
    projectsCount = len(projects)
    leadsDougnutLabels = ['Total Blue Badges', 'Leads', 'Capable Leads']
    leadsDougnutCount = [totalBlueCount,blueLeadsCount,leadsCapableCount]
    leadsDougnutLabels = json.dumps(leadsDougnutLabels)
    leadsDougnutCount = json.dumps(leadsDougnutCount)
    
    noProjDougnutLabels = ['Total Blue Badges', 'Blue Badges without Projects']
    noProjDougnutCount = [totalBlueCount,notAlignedToProjectsCount]
    noProjDougnutLabels = json.dumps(noProjDougnutLabels)
    noProjDougnutCount = json.dumps(noProjDougnutCount)

    return render(request,'redBlueEmpDetails.html',{'blueCount':totalBlueCount,'projectsCount':projectsCount,
        'projects':projects,'entCount':entCount,'spCount':spCount,'notAlignedToProjects':notAlignedToProjects,
        'blueLeads':blueLeads,'leadsCapable':leadsCapable,'leadsDougnutLabels':leadsDougnutLabels,
        'leadsDougnutCount':leadsDougnutCount,'noProjDougnutLabels':noProjDougnutLabels,'noProjDougnutCount':
        noProjDougnutCount})

def redResources(request):

    totalRedCount = Resources.objects.filter(badge='Red').count()
    redProjects = Resources.objects.filter(badge='Red')
    tempProj = redProjects.order_by().values('assignedProject').distinct()
    entCount = redProjects.filter(entOrSp='ENT').count()
    spCount = redProjects.filter(entOrSp='SP').count()
    notAlignedToProjects = redProjects.filter(assignedProject = 'Not Assigned')
    notAlignedToProjectsCount = notAlignedToProjects.count()
    redLeads = redProjects.filter(projectRole = 'Lead')
    leadsCapable = redProjects.filter(leadCapable = 'Yes')

    redLeadsCount = redLeads.count()
    leadsCapableCount = leadsCapable.count()

    projects = list()
    for pr in tempProj:
        if pr['assignedProject'] == 'Not Assigned':
            continue
        projects.append(pr['assignedProject'])

    projectsCount = len(projects)

    leadsDougnutLabels = ['Total Red Badges', 'Leads', 'Capable Leads']
    leadsDougnutCount = [totalRedCount,redLeadsCount,leadsCapableCount]
    leadsDougnutLabels = json.dumps(leadsDougnutLabels)
    leadsDougnutCount = json.dumps(leadsDougnutCount)

    noProjDougnutLabels = ['Total Red Badges', 'Red Badges without Projects']
    noProjDougnutCount = [totalRedCount,notAlignedToProjectsCount]
    noProjDougnutLabels = json.dumps(noProjDougnutLabels)
    noProjDougnutCount = json.dumps(noProjDougnutCount)

    return render(request, 'redResources.html',{'totalRedCount':totalRedCount,'projectsCount':projectsCount,'projects':projects,
        'entCount':entCount,'spCount':spCount,'notAlignedToProjectsCount':notAlignedToProjectsCount,
        'notAlignedToProjects':notAlignedToProjects,'leadsDougnutLabels':leadsDougnutLabels,'leadsDougnutCount':
        leadsDougnutCount,'redLeads':redLeads,'leadsCapable':leadsCapable,'noProjDougnutLabels':noProjDougnutLabels,
        'noProjDougnutCount':noProjDougnutCount})

def entTechnologies(request):

    availableresources = Resources.objects.filter(assignedProject = 'Not Assigned')
    availableresourcesSkills = list()
    freeEmpSkillDetails = list()
    totalEntSkilledResource = EntTechnologies.objects.all().count()

    for resource in availableresources:
        #print(resource.cecID)
        if len(EntTechnologies.objects.filter(cecId= resource.cecID)) ==0:
            continue
        availableresourcesSkills.append(EntTechnologies.objects.filter(cecId= resource.cecID).values('entRouting','entSwitching',
            'itilFoundation','entOperationsTools','entDesign','entDnaSDA','entMulticast','entQos','entDmvpn','entMeraki',
            'entMultiCloud','entSdn','entIpv6','cecId_id'))
    
    #print(list(availableresourcesSkills))
    names = list()
    for skill in availableresourcesSkills:
        freeEmpSkillDetails.append(skill[0])
        temp = Resources.objects.filter(cecID = skill[0]['cecId_id']).values('name')
        names.append(temp[0]['name'])
    
    labels = ['Total Ent Skilled Resources', 'Ent Resources Currently Free']
    counts = [totalEntSkilledResource,len(freeEmpSkillDetails)]

    labels = json.dumps(labels)
    counts = json.dumps(counts)

    allEntData = EntTechnologies.objects.all()

    return render(request,'EntTechnology.html',{'availableresourcesSkills':freeEmpSkillDetails,
        'freeEmpCount':len(freeEmpSkillDetails),'names':names,'labels':labels,'counts':counts})

def spTechnologies(request):

    availableresources = Resources.objects.filter(assignedProject = 'Not Assigned')
    availableresourcesSkills = list()
    freeEmpSkillDetails = list()
    totalSpSkilledResource = SpTechnologies.objects.all().count()

    for resource in availableresources:
        #print(resource.cecID)
        if len(SpTechnologies.objects.filter(cecId= resource.cecID)) ==0:
            continue
        availableresourcesSkills.append(SpTechnologies.objects.filter(cecId= resource.cecID).values('spRoutingSwitching','spMplsL2L3Vpn',
            'spMplste','spMulticastMultiCastVpn','spXrHardwarePlatformsOs','spNexusHardwarePlatformsOs','spAccessPlatforms',
            'spSegmentRouting','spSegmentRoutingSdnBgpPcep','spSegmentRoutingTE','spEvpnPbBEvpn','spVxlan','spNso','cecId_id'))

    names = list()
    for skill in availableresourcesSkills:
        freeEmpSkillDetails.append(skill[0])
        temp = Resources.objects.filter(cecID = skill[0]['cecId_id']).values('name')
        names.append(temp[0]['name'])

    labels = ['Total SP Skilled Resources', 'SP Resources Currently Free']
    counts = [totalSpSkilledResource,len(freeEmpSkillDetails)]
    
    labels = json.dumps(labels)
    counts = json.dumps(counts)
   

    return render(request,'SpTechnology.html',{'availableresourcesSkills':freeEmpSkillDetails,
        'freeEmpCount':len(freeEmpSkillDetails),'names':names,'labels':labels,'counts':counts})

def autTechnologies(request):
    
    scriptEmpData = ScriptingSkills.objects.all().values('python','vbaScripting','restApi','jsonXml','netconfRestconfYang',
        'cecId_id')
    
    entCounter = 0
    spCounter = 0
    names = list()
    empSkills = list()

    for emp in scriptEmpData:
        temp = Resources.objects.filter(cecID = emp['cecId_id']).values('entOrSp','name')
        if temp[0]['entOrSp'] == 'ENT':
            entCounter+=1
        else:
            spCounter+=1
        names.append(temp[0]['name'])
        empSkills.append(emp)
        #print(temp[0]['name'])
        #print(empSkills)

        labels = ['Resources In ENT Projects','Resources In SP Projects']
        counts = [entCounter,spCounter]

        labels = json.dumps(labels)
        counts = json.dumps(counts)

    return render(request,'Automation.html',{'empSkills':empSkills,'entCounter':entCounter,'spCounter':spCounter,
        'names':names,'scriptEmpDataCount':scriptEmpData.count(),'labels':labels,'counts':counts})


def projAutomation(request):

    availableProjects = Automation1.objects.all().values('projectName_id','offer','amiFunction','architecture','subDomain',
        'technology','qtr','lifeCycleStage','serviceDescription','solutionTemplateDocument','frequencyofdelivery',
        'effortBaseline','nwDiagUpdatesAutomatable','mwPrePostChecksAutomatable','configPushAutomatable',
        'validationCorporateAppAutomatable','validationPublicAppAutomatable','nwDiagUpdatesCurrentManualEffortMin',
        'mwPrePostChecksCurrentManualEffortMin','configPushCurrentManualEffortMin','validationCorporateAppCurrentManualEffortMin',
        'validationPublicAppCurrentManualEffortMin','nwDiagUpdatesCurrentFinalEffortMin',
        'mwPrePostChecksCurrentFinalEffortMin','configPushCurrentFinalEffortMin','validationCorporateAppFinalManualEffortMin',
        'validationPublicAppFinalManualEffortMin','nwDiagUpdatesDeliverableCount','mwPrePostChecksDeliverableCount',
        'configPushDeliverableCount','validationCorporateAppDeliverableCount','validationPublicAppDeliverableCount')

    projectsAutomated = 0
    AutomationProgress = list()
    averageMigration = list()
    projectName = list()
    namesWithNoAutomation = list()
    taskListValues = list()
    for project in availableProjects:

        if project['nwDiagUpdatesAutomatable'] == 'No'and project['mwPrePostChecksAutomatable'] == 'No' and project['configPushAutomatable'] == 'No' and project['validationCorporateAppAutomatable'] == 'No'and project['validationPublicAppAutomatable'] == 'No':
            namesWithNoAutomation.append(project['projectName_id'])
            continue

        else:
            nwDiagUpdatesFinal =  project['nwDiagUpdatesCurrentFinalEffortMin']
            mwPrePostChecksFinal = project['mwPrePostChecksCurrentFinalEffortMin']
            configPushCurrentFinal = project['configPushCurrentFinalEffortMin']
            validationCorporateAppFinal = project['validationCorporateAppFinalManualEffortMin']
            validationPublicAppFinal = project['validationPublicAppFinalManualEffortMin']

            nwDiagUpdatesManual = project['nwDiagUpdatesCurrentManualEffortMin']
            mwPrePostChecksManual = project['mwPrePostChecksCurrentManualEffortMin']
            configPushManual = project['configPushCurrentManualEffortMin']
            validationCorporateAppManual = project['validationCorporateAppCurrentManualEffortMin']
            validationPublicAppManual = project['validationPublicAppCurrentManualEffortMin']

            diagUpdates = (nwDiagUpdatesFinal/float(nwDiagUpdatesManual)) *100
            mwPrePost = (mwPrePostChecksFinal/float(mwPrePostChecksManual)) *100
            configPush = (configPushCurrentFinal/float(configPushManual)) *100
            validationCorporate = (validationCorporateAppFinal/float(validationCorporateAppManual)) *100
            validationPublic = (validationPublicAppFinal/float(validationPublicAppManual)) *100


            temp = [diagUpdates,mwPrePost,configPush,validationCorporate,validationPublic]
            taskListValues.append(temp)
            averageMigration.append(sum(temp)/len(temp))
            projectName.append(project['projectName_id'])
            AutomationProgress.append(temp)
            projectsAutomated+=1

    averageMigration = [round(num,1) for num in averageMigration]
    averagedMigrationValues = list()

    if len(taskListValues) > 1:
        temp1 = 0
        temp2 = 0
        temp3 = 0
        temp4 = 0
        temp5 = 0
        for projectsVal in taskListValues:
            temp1 += projectsVal[0]
            temp2 += projectsVal[1]
            temp3 += projectsVal[2]
            temp4 += projectsVal[3]
            temp5 += projectsVal[4]
        
        averagedMigrationValues=[temp1, temp2, temp3, temp4, temp5]
    else:
        averagedMigrationValues = taskListValues

    averagedMigrationValues = [num/len(taskListValues) for num in averagedMigrationValues]


    averagedMigrationValues = [round(num,1) for num in averagedMigrationValues]
    taskListLabel = ['NW Diagram Updates','NW PrePost Checks','config Push','Corporate Apps Validation','Public Cloud Apps Validation']

    zippedProjectInfo = zip(projectName,averageMigration)

    labels = ['Total Projects', 'Automated Projects']
    projectsCount = [availableProjects.count(),projectsAutomated]

    AutomationProgress = json.dumps(AutomationProgress)
    averageMigration = json.dumps(averageMigration)
    projectName = json.dumps(projectName)
    labels = json.dumps(labels)
    projectsCount = json.dumps(projectsCount)
    averagedMigrationValues = json.dumps(averagedMigrationValues)
    taskListLabel = json.dumps(taskListLabel)

    
    return render(request, 'ProjectAutomation.html',{'availableProjectsCount':availableProjects.count(),'projectsAutomated':
        projectsAutomated,'projectName':projectName,'averageMigration':averageMigration,'zippedProjectInfo':zippedProjectInfo,
        'namesWithNoAutomation':namesWithNoAutomation,'labels':labels,'projectsCount':projectsCount,'averagedMigrationValues':
        averagedMigrationValues,'taskListLabel':taskListLabel})


def projectInfo(request):

    projectsInfo = Resources.objects.order_by().values('assignedProject').distinct()

    projects = list()
    resourceCount = list()
    notAssigned = 0
    totalAssigned = 0
    projectsAll = list()
    projectsAllResources = list()

    for project in projectsInfo:

        if project['assignedProject'] == 'Not Assigned':
            notAssigned = Resources.objects.filter(assignedProject = project['assignedProject']).count()
            #projects.append(project['assignedProject'])
            #resourceCount.append(Resources.objects.filter(assignedProject = project['assignedProject']).count())
        else:
            totalAssigned += Resources.objects.filter(assignedProject = project['assignedProject']).count()
            projects.append(project['assignedProject'])
            resourceCount.append(Resources.objects.filter(assignedProject = project['assignedProject']).count())
        projectsAll.append(project['assignedProject'])
        projectsAllResources.append(Resources.objects.filter(assignedProject = project['assignedProject']).count())

    assignedNotAssignedLabel = ['Total Resources', 'Assigned To Projects', 'Not Assigned To Projects']
    assignedNotAssignedCount = [totalAssigned+notAssigned, totalAssigned ,notAssigned]

    projectsNames = projects

    projectsInfo = zip(projectsNames,resourceCount)
    projects = json.dumps(projects)
    #resourceCount = json.dumps(resourceCount)
    assignedNotAssignedLabel = json.dumps(assignedNotAssignedLabel)
    assignedNotAssignedCount = json.dumps(assignedNotAssignedCount)
    projectsAllResources = json.dumps(projectsAllResources)
    projectsAll = json.dumps(projectsAll)

    return render(request,'projectsInfoPage.html',{'assignedNotAssignedLabel':assignedNotAssignedLabel,
        'assignedNotAssignedCount':assignedNotAssignedCount,'projectsInfo':projectsInfo,'notAssigned':notAssigned,
        'projects':projects,'projectsAllResources':projectsAllResources,'projectsAll':projectsAll,
        'projectsCounts':len(projectsNames)})

def managersInfo(request):
    managers = Resources.objects.order_by().values('manager').distinct()
    
    manager = list()
    empCount = list()
    redCount = list()
    blueCount = list()

    for m in managers:
        resourceObj = Resources.objects.filter(manager = m['manager'])
        manager.append(m['manager'])
        empCount.append(resourceObj.count())
        redCount.append(resourceObj.filter(badge = 'Red').count())
        blueCount.append(resourceObj.filter(badge = 'Blue').count())

  
    totalEmp = sum(empCount)
    zipManagerCount = zip(manager,empCount,blueCount,redCount)
    manager = json.dumps(manager)
    redCount = json.dumps(redCount)
    blueCount = json.dumps(blueCount)
    empCount = json.dumps(empCount)

    return render(request,'managersInfo.html',{'zipManagerCount':zipManagerCount,'managerCount':managers.count(),
        'totalEmp': totalEmp,'manager':manager,'redCount':redCount,'blueCount':blueCount,'empCount':empCount})

def individualResources(request):
    resources = Resources.objects.all()
    return render(request,'IndividualResources.html',{'resources':resources})

def resourceDetails(request, cecID,entOrSp):

    name = list(Resources.objects.filter(cecID = cecID).values("name")[0].values())
    manager = list(Resources.objects.filter(cecID = cecID).values("manager")[0].values())
    project = list(Resources.objects.filter(cecID = cecID).values("assignedProject")[0].values())
    badge = list(Resources.objects.filter(cecID = cecID).values("badge")[0].values())
    role = list(Resources.objects.filter(cecID = cecID).values("projectRole")[0].values())
    
    certs = resourceSkills.objects.filter(cecId_id = cecID).values("certiDevnet","certiCcna","certiCcnp","certiSdwan",
        "certiSda","certiCcie")

    certsFlag = "true"
    entspFlag = ""
    skillsList = list()
    skillsLabel = list()

    skillsLabelEnt = ['Routing','Switching','ItilFoundation','OperationsTools','Design','DnaSDA','Multicast','Qos','Dmvpn',
    'Meraki','MultiCloud','Sdn','Ipv6']
    skillslabelSp = ['RoutingSwitching','MplsL2L3Vpn',
            'Mplste','MulticastMultiCastVpn','XrHardwarePlatformsOs','NexusHardwarePlatformsOs','AccessPlatforms',
            'SegmentRouting','SegmentRoutingSdnBgpPcep','SegmentRoutingTE','EvpnPbBEvpn','Vxlan','Nso']

    if entOrSp == "ENT":
        entspFlag = entOrSp
        availableresourcesSkills = EntTechnologies.objects.filter(cecId= cecID).values('entRouting','entSwitching',
            'itilFoundation','entOperationsTools','entDesign','entDnaSDA','entMulticast','entQos','entDmvpn','entMeraki',
            'entMultiCloud','entSdn','entIpv6','cecId_id')
        skills = list(availableresourcesSkills[0].values())
        for skill in skills:
            if 'Lead' in skill:
                skillsList.append(3)
            elif 'Specialist' in skill:
                skillsList.append(1)
            elif 'SME' in skill:
                skillsList.append(2)
            elif 'Completed' in skill:
                skillsList.append(1)
            else:
                skillsList.append(0)

        skillsLabel = skillsLabelEnt
    else:
        availableresourcesSkills = SpTechnologies.objects.filter(cecId= cecID).values('spRoutingSwitching','spMplsL2L3Vpn',
            'spMplste','spMulticastMultiCastVpn','spXrHardwarePlatformsOs','spNexusHardwarePlatformsOs','spAccessPlatforms',
            'spSegmentRouting','spSegmentRoutingSdnBgpPcep','spSegmentRoutingTE','spEvpnPbBEvpn','spVxlan','spNso','cecId_id')
        skills = list(availableresourcesSkills[0].values())
        for skill in skills:
            if 'Lead' in skill:
                skillsList.append(3)
            elif 'Specialist' in skill:
                skillsList.append(1)
            elif 'SME' in skill:
                skillsList.append(2)
            elif 'Completed' in skill:
                skillsList.append(1)
            else:
                skillsList.append(0)

        skillsLabel = skillslabelSp
    

    

    #skillsLabelEnt = json.dumps(skillsLabelEnt)
    skillsLabel = json.dumps(skillsLabel)
    skillsList = json.dumps(skillsList)
    #print(skillsList)
    
    #if no data is available in resourceskills table on the resource
    if len(certs) == 0:
        certsFlag = ""

        print(skillsLabel)
        print(skillsList)
        return render(request,'resourceDetails.html',{'name':name[0],'manager':manager[0],'project':project[0],'badge':badge[0],
            'role':role[0],'entOrSp':entOrSp,'certsFlag':certsFlag,'entspFlag':entspFlag,
            'availableresourcesSkills':availableresourcesSkills,'skillsLabel':skillsLabel,'skillsList':skillsList})

    certValue = list()
    certLabel = ['Devnet','CCNA','CCNP','SDWAN','SDA','CCIE']
    
    
    certValues = list(certs[0].values())

    for value in certValues:
        if value == 'Yes':
            certValue.append(1)
        else:
            certValue.append(0)

    certLabel = json.dumps(certLabel)
    certValue = json.dumps(certValue)
    certValues = json.dumps(certValues)
    
    return render(request,'resourceDetails.html',{'certLabel':certLabel,'certValue':certValue,'name':name[0],
        'certValues':certValues,'manager':manager[0],'project':project[0],'badge':badge[0],'role':role[0],
        'entOrSp':entOrSp,'certsFlag':certsFlag,'entspFlag':entspFlag,'availableresourcesSkills':availableresourcesSkills,
        'skillsLabel':skillsLabel,'skillsList':skillsList})

'''
def skillsSearch(request):
    if request.method == "POST":
    
        skills = request.POST['skills']
        skillsLevel = request.POST['skillsLevel']
        if skillsLevel == '1':
            skillsLevel = 'Specialist (Basic knowledge and hands on)'
        elif skillsLevel == '2':
            skillsLevel = 'SME (Can work independently)'
        else:
            skillsLevel = 'Lead (Can train others)'

        entSpFlag = ""
        if len(skills) == 0:
            return render(request, 'skillsSearchPage.html')
        else:
            entSpFlag = "true"

        
        entSp = ""
        skillsSplit = skills.split(';')
        columns = ""
        skillsList = list()



        for i in range(1, len(skillsSplit)):
            entSp = skillsSplit[0]
            columns += skillsSplit[0]+skillsSplit[i].strip()
            skillsList.append(skillsSplit[0]+skillsSplit[i].strip())
            if i == len(skillsSplit)-1:
                columns += ' = '+"\'"+skillsLevel+"\'"
                continue
            else:
                columns += ' = '+"\'"+skillsLevel+"\'"+' and '

        if entSp == "ent":
            query = ("SELECT cecId_id,name,badge,assignedProject,manager,cecId,entOrSp FROM EntTechnologies,Resources WHERE cecId_id = cecID and "
            + columns+";")
        else:
            query = ("SELECT cecId_id,name,badge,assignedProject,manager,cecId,entOrSp FROM SpTechnologies,Resources WHERE cecId_id = cecID and "
            + columns+";")
        print(query)
        val = selectDB(query)
        #print("***********")
        print(val)
        #print(val.values())
        selectedValues = list()
        for cols in val.values():
            selectedValues.append(cols)

        print(selectedValues)
        return render(request, 'skillsSearchPage.html',{'selectedValues':selectedValues,'entSpFlag':entSpFlag})
    
    return render(request, 'skillsSearchPage.html')

'''

def skillsSearch(request):
    if request.method == "POST":
    
        skills = request.POST['skills']
        skillsLevel = request.POST['skillsLevel']
        if skillsLevel == '1':
            skillsLevel = 'Specialist (Basic knowledge and hands on)'
        elif skillsLevel == '2':
            skillsLevel = 'SME (Can work independently)'
        else:
            skillsLevel = 'Lead (Can train others)'

        entSpFlag = ""
        if len(skills) == 0:
            return render(request, 'skillsSearchPage.html')
        else:
            entSpFlag = "true"

        
        entSp = ""
        skillsSplit = skills.split(',')
        columns = ""
        skillsList = list()



        #selectskills = ("SELECT  FROM projectdetails2  WHERE projectName = \'"+str(projectName)+"\';")
        
        for i in range(0, len(skillsSplit)):
            #entSp = skillsSplit[0]
            columns += skillsSplit[i].strip()
            skillsList.append(skillsSplit[0]+skillsSplit[i].strip())
            if i == len(skillsSplit)-1:
                columns += ' = '+"\'"+skillsLevel+"\'"
                continue
            else:
                columns += ' = '+"\'"+skillsLevel+"\'"+' and '

        
        query = ("SELECT r.cecID, r.name, r.assignedProject, r.manager, r.entOrSp FROM enttechnologies e, sptechnologies s, Resources r WHERE e.cecId = s.cecId and r.cecID = e.cecId and r.cecID = s.cecId and "
            + columns+";")

        print(query)
        val = selectDB(query)
        #print("***********")
        print(val)
        #print(val.values())
        selectedValues = list()
        for k,cols in val.items():
            selectedValues.append(k)
            for col in cols:
                selectedValues.append(col)

        print(selectedValues)
        selectedValues = [selectedValues]
        return render(request, 'skillsSearchPage.html',{'selectedValues':selectedValues,'entSpFlag':entSpFlag})
    
    return render(request, 'skillsSearchPage.html')

def uploadInFlightProjectDetails(request):
    if "GET" == request.method:
        return render(request, 'uploadProjectDetails.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Migration In-Flight Project Tra"]
        #worksheet2 = wb["Sheet2"]
        #print(worksheet)

        excel_data = list()
        excel_data_sheet2 = list()
        # iterating over the rows and
        # getting value from each cell in row
        val = 0
        for i,row in enumerate(worksheet.iter_rows()):
            #lessons = ProjectDetails2.objects.all()
            #print(lessons.count())
            row_data = list()
            #row_data.append(lessons.count()+1)
            if i == 0:
                continue
            for cell in row:
                row_data.append((cell.value))
            excel_data.append(row_data)
            
            tempQuery = ("SELECT projectName from projectdetails2 where projectName = \'"+str(row_data[0])+"\';")
            rows = listDB(tempQuery)
            deletedRows = 0
            if(len(rows) >= 1):
                deleteQuery = ("DELETE from projectdetails2 where projectName = \'"+str(row_data[0])+"\';")
                deletedRows = deleteDB(deleteQuery)
                updateParentQuery = ("UPDATE Dashboard_Info set category = \'"+str(row_data[4])+"\', dmName = \'"+str(row_data[23])+"\', completionZone = \'"+str(row_data[14])+"\', projectStatus = \'"+str(row_data[14])+"\' WHERE projectName = \'"+str(row_data[0])+"\';")
                rowCount = updateDB(updateParentQuery)
                print("Updated "+str(rowCount)+ "rows")

                print("Deleted repeated entry with Project Name: "+str(row_data[0]))
            else:
                tempList = list()
                tempList.append(row_data[0])
                tempList.append(row_data[4])
                tempList.append(row_data[23])
                tempList.append(row_data[14])
                tempList.append(row_data[14])
                tempList.append(row_data[52])
                temp2 = ("INSERT INTO Dashboard_Info (projectName,category,dmName,completionZone,projectStatus,dealValue) values (%s,%s,%s,%s,%s,%s);")
                insertDB(temp2,tempList)

            

            query = ("INSERT INTO projectdetails2 (projectName,endDate,India_NCEs,otherRegionNCEs,category,technology,lead1,ProjectId,projectType,projectClass,preSalesActivity,preSalesContributor,preSalesEffort,customerSentiment,overallStatus,scheduleStatus,qualityStatus,resourceStatus,automationStatus,cbuCfuStatus,CommentsOnScheduleResourcing,executiveSummary,generalComments,projectDeliveryManager,status1,projectCompletionLevel,mwSuccessPercent,CovidUpdateTrigger,BillableWorkImpactedduetoCovid19,ImpactType,impactsummaryandinfo,TotalnumberofNCEassigned,NCEfilingBAUhours,BAUNormalCBUAllocation,NCEfilingreducedhours,ReducedCBUAllocation,NCEfilingZerohours,ImpactedNoofWeeks,AverageProjectCBU,TotalNCEs,AvgProjectCBUChange,Forhowmanyweeks,region,projectManager,projectLead,serviceDeliveryManager,source1,dealId,includeForBookings,bookingsComment,sourceRequesterName,staffing,dealValue,modifiedBy,repeatBusiness,repeatBusinessRemark,startDate,newProjectAudit,pOneProject,newProjectAuditDate,pOneAudit,pOneAuditDate,modified,test) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);")
            val = insertDB(query,row_data)
            if val:
                val+=1
        '''    
        if val:
                messages.success(request, "Successfully Uploaded")
                #return render(request, 'uploadProjectDetails.html', {"excel_data":excel_data})
        else:
                messages.error(request, "Monthly Data ProjectId already exists!, please try again")
        '''
        messages.success(request, "Successfully Uploaded")
    
        return render(request, 'uploadProjectDetails.html', {"excel_data":excel_data})
   
'''       
def uploadInMasterProjectsDB(request):
    if "GET" == request.method:
        return render(request, 'uploadMasterProjectDB.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Automated Migrations Master Pro"]
        #worksheet2 = wb["Sheet2"]
        #print(worksheet)

        excel_data = list()
        excel_data_sheet2 = list()
        # iterating over the rows and
        # getting value from each cell in row
        val = 0
        for i,row in enumerate(worksheet.iter_rows()):
            #lessons = ProjectDetails2.objects.all()
            #print(lessons.count())
            row_data = list()
            #row_data.append(lessons.count()+1)
            if i == 0:
                continue
            for cell in row:
                row_data.append((cell.value))
            excel_data.append(row_data)
            
            tempQuery = ("SELECT projectName from migration_Projects where projectName = \'"+str(row_data[1])+"\';")
            rows = listDB(tempQuery)
            deletedRows = 0
            if(len(rows) >= 1):
                deleteQuery = ("DELETE from migration_Projects where projectName = \'"+str(row_data[1])+"\';")
                deletedRows = deleteDB(deleteQuery)

            print("Deleted repeated entry with Project Name: "+str(row_data[1]))
            #print(len(row_data))
            query = ("INSERT INTO migration_Projects values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);")
            
            dashSelectQuery = ("SELECT projectName from Dashboard_Info where projectName = \'"+str(row_data[1])+"\';")
            
            if len(dashSelectQuery) >=1 :
                updateParentQuery = ("UPDATE Dashboard_Info set dealValue = \'"+str(row_data[14])+"\' where projectName = \'"+str(row_data[1])+"\';")
                rowCount = updateDB(updateParentQuery)
            else:
                tempList = list()
                tempList.append(row_data[1])
                tempList.append(row_data[6])
                tempList.append("")
                tempList.append("")
                tempList.append("")
                tempList.append(row_data[14])
                temp2 = ("INSERT INTO Dashboard_Info (projectName,category,dmName,completionZone,projectStatus,dealValue) values (%s,%s,%s,%s,%s,%s);")
                insertDB(temp2,tempList)
            val = insertDB(query,row_data)
            if val:
                val+=1
            updateParentQuery = ("UPDATE Dashboard_Info set dealValue = \'"+str(row_data[14])+"\' where projectName = \'"+str(row_data[1])+"\';")
            rowCount = updateDB(updateParentQuery)
            
        if val:
                messages.success(request, "Successfully Uploaded")
        else:
                messages.error(request, "Monthly Data ProjectId already exists!, please try again")
        
        return render(request, 'uploadMasterProjectDB.html', {"excel_data":excel_data})
'''

def uploadInMasterProjectsDB(request):
    if "GET" == request.method:
        return render(request, 'uploadMasterProjectDB.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Automated Migrations Master Pro"]
        #worksheet2 = wb["Sheet2"]
        #print(worksheet)

        excel_data = list()
        excel_data_sheet2 = list()
        # iterating over the rows and
        # getting value from each cell in row
        val = 0
        for i,row in enumerate(worksheet.iter_rows()):
            #lessons = ProjectDetails2.objects.all()
            #print(lessons.count())
            row_data = list()
            #row_data.append(lessons.count()+1)
            if i == 0:
                continue
            for cell in row:
                row_data.append((cell.value))
            excel_data.append(row_data)
            
            tempQuery = ("SELECT projectName from migration_Projects where projectName = \'"+str(row_data[1])+"\';")
            rows = listDB(tempQuery)
            deletedRows = 0
            if(len(rows) >= 1):
                deleteQuery = ("DELETE from migration_Projects where projectName = \'"+str(row_data[1])+"\';")
                deletedRows = deleteDB(deleteQuery)
                #delete from parent
                #deleteQuery1 = ("DELETE from Dashboard_Info where projectName = \'"+str(row_data[1])+"\';")
                #deletedRows = deleteDB(deleteQuery1)

            #print("Deleted repeated entry with Project Name: "+str(row_data[1]))
            #print(len(row_data))
            query = ("INSERT INTO migration_Projects values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);")

            dashSelectQuery = ("SELECT projectName from Dashboard_Info where projectName = \'"+str(row_data[1])+"\';")
            dashSelect = selectDB(dashSelectQuery)

            deal = row_data[14]
            #if len(row_data[14]) > 14:
                    #deal = "0"
            #else:
                    #deal = row_data[14]

            if len(dashSelect) >=1 :
                updateParentQuery = ("UPDATE Dashboard_Info set dealValue = \'"+str(deal)+"\' where projectName = \'"+str(row_data[1])+"\';")
                rowCount = updateDB(updateParentQuery)
            else:
                tempList = list()
                tempList.append(row_data[1])
                tempList.append(row_data[6])
                tempList.append("")
                tempList.append("")
                tempList.append("")
                tempList.append(deal)
                #tempList.append("")
                #tempList.append("")
                print("***********")
                print(tempList)
                #temp2 = ("INSERT INTO Dashboard_Info values (%s,%s,%s,%s,%s,%s,%s,%s);")
                temp2 = ("INSERT INTO Dashboard_Info(projectName,category,dmName,completionZone,projectStatus,dealValue) values (%s,%s,%s,%s,%s,%s);")
                val = insertDB(temp2,tempList)

            val = insertDB(query,row_data)
            
        messages.success(request, "Successfully Uploaded")
        
        return render(request, 'uploadMasterProjectDB.html', {"excel_data":excel_data})
'''
def uploadavailabilityTracker(request):
    if "GET" == request.method:
        return render(request, 'uploadAvailabilityTracker.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["EN Migrations - Availability Tr"]
        #worksheet2 = wb["Sheet2"]
        #print(worksheet)

        excel_data = list()
        excel_data_sheet2 = list()
        # iterating over the rows and
        # getting value from each cell in row
        val = 0
        for i,row in enumerate(worksheet.iter_rows()):
            row_data = list()
            if i == 0:
                continue
            if i <= 4:
                continue
            for cell in row:
                row_data.append((cell.value))
            excel_data.append(row_data)
            
            print(row_data[2])
            tempQuery = ("SELECT nce from availabilityTracker where nce = \'"+str(row_data[2])+"\';")
            rows = listDB(tempQuery)
            deletedRows = 0
            if(len(rows) >= 1):
                deleteQuery = ("DELETE from availabilityTracker where nce = \'"+str(row_data[2])+"\';")
                deletedRows = deleteDB(deleteQuery)

            print("Deleted repeated entry with NCE Name: "+str(row_data[2]))
            #print(len(row_data))
            query = ("INSERT INTO availabilityTracker values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);")
            val = insertDB(query,row_data)
            if val:
                val+=1
            #updateParentQuery = ("UPDATE Dashboard_Info set dealValue = \'"+str(row_data[14])+"\' where projectName = \'"+str(row_data[1])+"\';")
            #rowCount = updateDB(updateParentQuery)
            
        if val:
                messages.success(request, "Successfully Uploaded")
        else:
                messages.error(request, "Monthly Data ProjectId already exists!, please try again")
        
        messages.success(request, "Successfully Uploaded")
        return render(request, 'uploadAvailabilityTracker.html', {"excel_data":excel_data})

'''

def uploadavailabilityTracker(request):
    if "GET" == request.method:
        return render(request, 'uploadAvailabilityTracker.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["EN Migrations - Availability Tr"]
        #worksheet2 = wb["Sheet2"]
        #print(worksheet)

        excel_data = list()
        excel_data_sheet2 = list()
        # iterating over the rows and
        # getting value from each cell in row
        try:
            #val = 0
            for i,row in enumerate(worksheet.iter_rows()):
                row_data = list()
                if i == 0 or i == 1 or i == 2 or i == 3 or i == 4:
                    continue

                for cell in row:
                    row_data.append((cell.value))
                excel_data.append(row_data)
            
                #print(row_data[2])
                tempQuery = ("SELECT nce from availabilityTracker where nce = \'"+str(row_data[2])+"\';")
                rows = listDB(tempQuery)
                deletedRows = 0
                if(len(rows) >= 1):
                    deleteQuery = ("DELETE from availabilityTracker where nce = \'"+str(row_data[2])+"\';")
                    deletedRows = deleteDB(deleteQuery)

                #print("Deleted repeated entry with NCE Name: "+str(row_data[2]))
                print(len(row_data))
                print(row_data)
                query = ("INSERT INTO availabilityTracker values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);")
                val = insertDB(query,row_data)
                #if val:
                 #   val+=1
                #updateParentQuery = ("UPDATE Dashboard_Info set dealValue = \'"+str(row_data[14])+"\' where projectName = \'"+str(row_data[1])+"\';")
                #rowCount = updateDB(updateParentQuery)
            '''    
            if val:
                    messages.success(request, "Successfully Uploaded")
            else:
                    messages.error(request, "Monthly Data ProjectId already exists!, please try again")
            '''
        
            messages.success(request, "Successfully Uploaded")

            print(len(row_data))
            print(len(row))
            return render(request, 'uploadAvailabilityTracker.html', {"excel_data":excel_data})
        except:
            return HttpResponse("OOPS! Seems there is no data for the project searched")

def showAvailabilityTracker(request):
    selectProjQuery = ("SELECT nce,reportingManager,cec,role1,project,availabilityMonth1,availabilityMonth2,availabilityMonth3,availabilityMonth4,availabilityMonth5 FROM availabilityTracker  WHERE CHAR_LENGTH(nce)>2;")
    selectManagerQuery = ("SELECT reportingManager, nce FROM availabilityTracker  WHERE CHAR_LENGTH(nce)<=2;")

    result = selectDB(selectProjQuery) 
    managerRes = selectDB(selectManagerQuery)
    currentMonth = datetime.now().month
    #currentMonth = 9
    months = ['Jan','Feb','Mar','Apr','May','June','July','Aug','Sep','Oct','Nov','Dec']
    
    
    
    monthsTracker = list()
    months1AvailabilityTracker = list()
    months2AvailabilityTracker = list()
    months3AvailabilityTracker = list()
    months4AvailabilityTracker = list()
    months5AvailabilityTracker = list()

    if currentMonth <= 8:
        for i in range(currentMonth-1,currentMonth+4):
            monthsTracker.append(months[i])
    else:
        for i in range(currentMonth-1,currentMonth+4):
            monthsTracker.append(months[i%len(months)])

    mon11 = monthsTracker[0]
    mon22 = monthsTracker[1]
    mon33 = monthsTracker[2]
    mon44 = monthsTracker[3]
    mon55 = monthsTracker[4]
    
    nceList = list()
    managerList = list()
    cecList = list()
    roleList = list()
    projectList = list()
    managersList = list()
    empCount = list()

    for k,v in managerRes.items():
        managersList.append(k)
        empCount.append(v[0])

    print(managersList)
    print(empCount)
    for k,v in result.items():
        tempList = list()
        tempList.append(k)
        nceList.append(k)
        managerList.append(v[0])
        #print(v)
        cecList.append(v[1])
        roleList.append(v[2])
        projectList.append(v[3])
        
        if v[4] is None:
            mon1 = 0
        else:
            mon1 = float(v[4]) * 100
            
        if v[5] is None:
            mon2 = 0
        else:
            mon2 = float(v[5]) * 100
            
        if v[6] is None:
            mon3 = 0
        else:
            mon3 = float(v[6]) * 100
            
        if v[7] is None:
            mon4 = 0
        else:
            mon4 = float(v[7]) * 100
            
        if v[8] is None:
            mon5 = 0
        else:
            mon5 = float(v[8]) * 100
            
        dummyList = list()
        for i in range(0 ,len(nceList)):
            dummyList.append("")
            
        months1AvailabilityTracker.append(mon1)
        months2AvailabilityTracker.append(mon2)
        months3AvailabilityTracker.append(mon3)
        months4AvailabilityTracker.append(mon4)
        months5AvailabilityTracker.append(mon5)
        
        '''
        months1AvailabilityTracker.append(monthsTracker[0]+" & "+str(v[4])+"%")
        months2AvailabilityTracker.append(monthsTracker[1]+" & "+str(v[4+1])+"%")
        months3AvailabilityTracker.append(monthsTracker[2]+" & "+str(v[4+2])+"%")
        months4AvailabilityTracker.append(monthsTracker[3]+" & "+str(v[4+3])+"%")
        months5AvailabilityTracker.append(monthsTracker[4]+" & "+str(v[4+4])+"%")
        '''

    zippedList = zip(dummyList,nceList,managerList,cecList,roleList,projectList,months1AvailabilityTracker,months2AvailabilityTracker,months3AvailabilityTracker,months4AvailabilityTracker,months5AvailabilityTracker)
      
    managersList = json.dumps(managersList) 
    empCount = json.dumps(empCount) 

    return render(request, 'showAvailabilityTracker.html',{'zippedList':zippedList,'managersList':managersList,'empCount':empCount,'mon11':mon11,'mon22':mon22,'mon33':mon33,'mon44':mon44,'mon55':mon55})

def showprojectDetails_new(request, projectName):
    selectProjQuery = ("SELECT * FROM projectdetails2  WHERE projectName = \'"+str(projectName)+"\';")
    selectProjQuery2 = ("SELECT * FROM migration_Projects  WHERE projectName = \'"+str(projectName)+"\';")
    result = selectDBParam(selectProjQuery,projectName)
    result2 = selectDBParam(selectProjQuery2,projectName)

    try:
    
        projectData = result[projectName]
        data2 = result2[projectName]
        
        try:
            avgCBU = str(float(projectData[38])*100)
        except:
            avgCBU = "NA"
        region = projectData[42]
        try:
            wSuccess = str(float(data2[18])*100)+"%"
        except:
            wSuccess = "NA"
        
        try:
            cVal = returnStatusValue(projectData[13])
        except:
            cVal = 0
        try:
            oVal = returnStatusValue(projectData[14])
        except:
            oVal = 0
        try:
            aVal = returnStatusValue(projectData[18])
        except:
            aVal = 0
        try:
            sVal = returnStatusValue(projectData[15])
        except:
            sVal = 0
        try:
            qVal = returnStatusValue(projectData[16])
        except:
            qVal = 0
        try:
            migrateDevices = data2[16]
        except:
            migrateDevices = 0
        try:
            executedWindows = data2[17]
        except:
            executedWindows = 0
        try:
            totalNCE = projectData[39]
        except:
            totalNCE = 0
        try:
            averageCBU = float(projectData[38])*100
        except:
            averageCBU = "NA"

        dataList = list()
        statusList = list()
        statusLabels = ['Customer Sentiment','Overall Status','Schedule Status','Quality Status','Resources Status','Automation Status','CBU/CFU Status']
        nceInfo = list()

        dataList.append(migrateDevices)
        dataList.append(executedWindows)

        for i in range(13,20):
            if projectData[i] == 'Green':
                statusList.append(3)
            elif projectData[i] == 'Red':
                statusList.append(1)
            else:
                statusList.append(2)

        for i in range(32,40):
            try:
                nceInfo.append(projectData[i])
            except:
                nceInfo.append(" ")

        dataList  = json.dumps(dataList)
        statusList = json.dumps(statusList)
        statusLabels = json.dumps(statusLabels)

        return render(request, 'projectDetails.html', {'numberOfSitesPlanned':65,'numOfSitesMigrated':40,'pID': projectData[7], 'projectName':projectName, 'summary':projectData[21], 'nceIndia':projectData[2], 'nceOther':projectData[3],'avgCBU':avgCBU, 'region': region, 'endDate': projectData[1], 'comments':projectData[22],'techno':projectData[5],'bookingYear':data2[1],'eFY':data2[2],'lead':data2[3],'mTrack':data2[6],'osVer':data2[7],'dMigrated':data2[16],'wExecuted':data2[17],'wSuccess':wSuccess,'pStatus':data2[19],'cVal':cVal,'custSenti':projectData[13],'oVal':oVal,'oStatus':projectData[14],'aVal':aVal,'aStatus': projectData[18], 'sStatus':projectData[15],'sVal':sVal,'qVal':qVal, 'qStatus':projectData[16],'dManager':projectData[23],'dStatus':projectData[24],'pCompletionLevel':projectData[25],'dataList':dataList,'statusLabels':statusLabels,'statusList':statusList,'totalNCE':totalNCE,'averageCBU':averageCBU})
    except:
        return HttpResponse("OOPS! Seems there is no data for the project searched")

def showprojectDetails_new_Temp(request, projectName):
    selectProjQuery = ("SELECT * FROM projectdetails2  WHERE projectName = \'"+str(projectName)+"\';")
    selectProjQuery2 = ("SELECT * FROM migration_Projects  WHERE projectName = \'"+str(projectName)+"\';")
    result = selectDBParam(selectProjQuery,projectName)
    result2 = selectDBParam(selectProjQuery2,projectName)

    
    print(result)
    projectData = result[projectName]
    data2 = result2[projectName]
    
    try:
        avgCBU = str(float(projectData[38])*100)
    except:
        avgCBU = "NA"
    region = projectData[42]
    try:
        wSuccess = str(float(data2[18])*100)+"%"
    except:
        wSuccess = "NA"
        
    try:
        cVal = returnStatusValue(projectData[13])
    except:
        cVal = 0
    try:
        oVal = returnStatusValue(projectData[14])
    except:
        oVal = 0
    try:
        aVal = returnStatusValue(projectData[18])
    except:
        aVal = 0
    try:
        sVal = returnStatusValue(projectData[15])
    except:
        sVal = 0
    try:
        qVal = returnStatusValue(projectData[16])
    except:
        qVal = 0
    try:
        migrateDevices = data2[16]
    except:
        migrateDevices = 0
    try:
        executedWindows = data2[17]
    except:
        executedWindows = 0
    try:
        totalNCE = projectData[39]
    except:
        totalNCE = 0
    try:
        averageCBU = float(projectData[38])*100
    except:
        averageCBU = "NA"

    dataList = list()
    statusList = list()
    statusLabels = ['Customer Sentiment','Overall Status','Schedule Status','Quality Status','Resources Status','Automation Status','CBU/CFU Status']
    nceInfo = list()

    dataList.append(migrateDevices)
    dataList.append(executedWindows)

    for i in range(13,20):
        if projectData[i] == 'Green':
            statusList.append(3)
        elif projectData[i] == 'Red':
            statusList.append(1)
        else:
            statusList.append(2)

    for i in range(32,40):
        try:
            nceInfo.append(projectData[i])
        except:
            nceInfo.append(" ")

    dataList  = json.dumps(dataList)
    statusList = json.dumps(statusList)
    statusLabels = json.dumps(statusLabels)

    return render(request, 'projectDetails.html', {'numberOfSitesPlanned':65,'numOfSitesMigrated':40,'pID': projectData[7], 'projectName':projectName, 'summary':projectData[21], 'nceIndia':projectData[2], 'nceOther':projectData[3],'avgCBU':avgCBU, 'region': region, 'endDate': projectData[1], 'comments':projectData[22],'techno':projectData[5],'bookingYear':data2[1],'eFY':data2[2],'lead':data2[3],'mTrack':data2[6],'osVer':data2[7],'dMigrated':data2[16],'wExecuted':data2[17],'wSuccess':wSuccess,'pStatus':data2[19],'cVal':cVal,'custSenti':projectData[13],'oVal':oVal,'oStatus':projectData[14],'aVal':aVal,'aStatus': projectData[18], 'sStatus':projectData[15],'sVal':sVal,'qVal':qVal, 'qStatus':projectData[16],'dManager':projectData[23],'dStatus':projectData[24],'pCompletionLevel':projectData[25],'dataList':dataList,'statusLabels':statusLabels,'statusList':statusList,'totalNCE':totalNCE,'averageCBU':averageCBU})
    
    '''
    for ragRows in projRagInfo:
        #year = datetime.datetime.strptime(str(ragRows),'%Y-%m')
        siteComp = siteComp + ragRows.sitesCompleted
        month = str(ragRows.month)
        year = str(ragRows.year)
        mli = year+'-'+month
        winCompletedList.append(ragRows.windowsCompleted)
        monthList.append(ragRows.month)

    if projDetInfo.count() != 0:
        completedPercent = round((siteComp/tSites)*100)
        sitesInProgress = tSites - siteComp
        sitesInProgress = round((sitesInProgress/tSites)*100)
        projDetails = ProjectDetails2.objects.get(ProjectId = projDetInfo[0].ProjectId)
        ragDetails = RagDetails.objects.filter(ProjectId = projDetInfo[0].ProjectId)
        completedSites = list()
        xAxis = list()
        custSatisfaction = list()
        windowsPlanned = list()
        windowsCompleted = list()
        
        for rag in ragDetails:
            completedSites.append(rag.sitesCompleted)
            xAxis.append(str(rag.year)+' Month '+str(rag.month))
            custSatisfaction.append(rag.custSatisfaction)
            windowsPlanned.append(rag.windowsPlanned)
            windowsCompleted.append(rag.windowsCompleted)
            
       
        #print(completedSites)
        #print(xAxis)
        completedSites = json.dumps(completedSites)
        custSatisfaction = json.dumps(custSatisfaction)
        xAxis = json.dumps(xAxis)

        #print(windowsPlanned)
        #print(windowsCompleted)

        windowSuccessRate = list()
        
        
        for i in range(len(windowsPlanned)):
            #print(float(windowsCompleted[i]))
            temp1 = float(windowsCompleted[i])
            #print(temp1)
            print(windowsPlanned[i])
            print(windowsCompleted[i])
            if windowsCompleted[i] > windowsPlanned[i]:
                windowSuccessRate.append(100)
            else:
                windowSuccessRate.append((temp1/windowsPlanned[i])*100)

            #print(temp1/windowsPlanned[i])
            #print((float64(windowsCompleted[i])//float64(windowsPlanned[i])))
            #print(windowsCompleted[i])

        

        windowsPlanned = json.dumps(windowsPlanned)
        windowsCompleted = json.dumps(windowsCompleted)
        windowSuccessRate = json.dumps(windowSuccessRate)
        
        #print(windowSuccessRate)
        #print(completedSites)
        #print(xAxis)
        #print(projDetInfo[0].ProjectId)
        
        

        #for i in range(len(windowsPlanned)):
         #   print(windowsPlanned[i])
          #  print(windowsCompleted[i])

        return render(request, 'projectDetails.html', {'pdInfo': projDetInfo[0], 'winCompletedList':winCompletedList,'monthList':monthList,
        'siteComp':siteComp, 'tSites':tSites,'projectName':projectName,
        'completedPercent':completedPercent,'sitesInProgress':sitesInProgress,
        'projDetails':projDetails,'ragDetails':ragDetails,'xAxis':xAxis,'completedSites':completedSites,
        'custSatisfaction':custSatisfaction,'windowsPlanned':windowsPlanned,
        'windowsCompleted':windowsCompleted,'windowSuccessRate':windowSuccessRate})
        #return render(request, 'rough.html', {'pdInfo': projDetInfo[0], 'winCompletedList':winCompletedList,'monthList':monthList,
        #'siteComp':siteComp, 'tSites':tSites})
    else:
        return HttpResponse("<h3> OOPS! The requested Project '"+projectName+"' does not have any data yet")
    '''
def uploadaMigrationTrainingDetails(request):
    if "GET" == request.method:
        return render(request, 'uploadTrainingTracker.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Baseline - Migration Training"]
        #worksheet2 = wb["Sheet2"]
        #print(worksheet)

        excel_data = list()
        excel_data_sheet2 = list()
        
        # iterating over the rows and
        # getting value from each cell in row
        val = 0
        for i,row in enumerate(worksheet.iter_rows()):
            resourceTable = list()
            if i==0:
                continue
            row_data = list()

            for cell in row:
                row_data.append((cell.value))
            excel_data.append(row_data)
            
            if row_data[0] is None:
                continue
            else:
                resourceTable.append(row_data[0].split("@")[0])
            resourceTable.append(row_data[1])
            resourceTable.append(0)
            resourceTable.append("NA")
            resourceTable.append("NA")
            resourceTable.append(row_data[0])
            resourceTable.append("NA")
            resourceTable.append("NA")
            resourceTable.append(row_data[2])
            resourceTable.append("NA")
            resourceTable.append(row_data[3])
            resourceTable.append(row_data[4])

            tempQuery = ("SELECT cecid from Resources where cecid = \'"+str(row_data[0].split("@")[0])+"\';")
            rows = listDB(tempQuery)
            deletedRows = 0

            #delete exsisting data from 2 children tables and 1 parent Resource table before updating new values
            if(len(rows) >= 1):
                deleteQuery1 = ("DELETE from enttechnologies where cecid = \'"+str(row_data[0].split("@")[0])+"\';")
                deletedRows = deleteDB(deleteQuery1)
                deleteQuery2 = ("DELETE from sptechnologies where cecid = \'"+str(row_data[0].split("@")[0])+"\';")
                deletedRows = deleteDB(deleteQuery2)
                deleteQuery3 = ("DELETE from migrationLifeCycle where cecid = \'"+str(row_data[0].split("@")[0])+"\';")
                deletedRows = deleteDB(deleteQuery3)
                deleteQuery4 = ("DELETE from scripting where cecid = \'"+str(row_data[0].split("@")[0])+"\';")
                deletedRows = deleteDB(deleteQuery4)
                deleteQuery5 = ("DELETE from activeccie where cecid = \'"+str(row_data[0].split("@")[0])+"\';")
                deletedRows = deleteDB(deleteQuery5)
                deleteQuery6 = ("DELETE from othercerts where cecid = \'"+str(row_data[0].split("@")[0])+"\';")
                deletedRows = deleteDB(deleteQuery6)
                deleteQuery7 = ("DELETE from competitiveplatforms where cecid = \'"+str(row_data[0].split("@")[0])+"\';")
                deletedRows = deleteDB(deleteQuery7)
                deleteQuery8 = ("DELETE from softskills where cecid = \'"+str(row_data[0].split("@")[0])+"\';")
                deletedRows = deleteDB(deleteQuery8)
                deleteQuery = ("DELETE from Resources where cecid = \'"+str(row_data[0].split("@")[0])+"\';")
                deletedRows = deleteDB(deleteQuery)
                print("Deleted repeated entry with Email: "+str(row_data[0]))


            insertQuery = ("INSERT INTO Resources values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);")
            insertQuery1 = ("INSERT INTO enttechnologies values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);")
            insertQuery2 = ("INSERT INTO sptechnologies values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);")
            insertQuery3 = ("INSERT INTO migrationLifeCycle values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);")
            insertQuery4 = ("INSERT INTO scripting values (%s,%s,%s,%s,%s,%s,%s);")
            insertQuery5 = ("INSERT INTO activeccie values (%s,%s,%s,%s,%s,%s,%s,%s,%s);")
            insertQuery6 = ("INSERT INTO othercerts values (%s,%s,%s,%s,%s);")
            insertQuery7 = ("INSERT INTO competitiveplatforms values (%s,%s,%s,%s,%s,%s,%s);")
            insertQuery8 = ("INSERT INTO softskills values (%s,%s,%s,%s,%s,%s);")

            entList = list()
            spList = list()
            lifeCycleList = list()
            scripting = list()
            activeCcie = list()
            othercertsList = list()
            competitiveplatforms = list()
            softskills = list()

            entList.append(row_data[0].split("@")[0])
            spList.append(row_data[0].split("@")[0])
            lifeCycleList.append(row_data[0].split("@")[0])
            scripting.append(row_data[0].split("@")[0])
            activeCcie.append(row_data[0].split("@")[0])
            othercertsList.append(row_data[0].split("@")[0])
            competitiveplatforms.append(row_data[0].split("@")[0])
            softskills.append(row_data[0].split("@")[0])

            for i in range(6,27):
                entList.append(row_data[i])
            for i in range(28,50):
                spList.append(row_data[i])
            for i in range(51,67):
                lifeCycleList.append(row_data[i])
            for i in range(68,74):
                scripting.append(row_data[i])
            for i in range(75,83):
                activeCcie.append(row_data[i])
            for i in range(84,88):
                othercertsList.append(row_data[i])
            for i in range(89,95):
                competitiveplatforms.append(row_data[i])
            for i in range(96,101):
                softskills.append(row_data[i])

            
            insertDB(insertQuery,resourceTable)
            insertDB(insertQuery1,entList)
            insertDB(insertQuery2,spList)
            insertDB(insertQuery3,lifeCycleList)
            insertDB(insertQuery4,scripting)
            insertDB(insertQuery5,activeCcie)
            insertDB(insertQuery6,othercertsList)
            insertDB(insertQuery7,competitiveplatforms)
            insertDB(insertQuery8,softskills)
        
        messages.success(request, "Successfully Uploaded")
        return render(request, 'uploadTrainingTracker.html', {"excel_data":excel_data})
        
def resourceDetails1(request, cecID):
    #cecID
    try:
        selectResourceQuery = ("SELECT cecId,name,email,assignedProject,entOrSp,manager,inclination FROM Resources  WHERE cecID = \'"+str(cecID)+"\';")
        selectCertQuery = ("SELECT cecId,devnet,devnetCertifiedDate,specialistEntSdWan,specialistEntSdWanDate FROM othercerts  WHERE cecID = \'"+str(cecID)+"\';")
        selectentQuery = ("SELECT * FROM enttechnologies  WHERE cecID = \'"+str(cecID)+"\';")
        selectSpQuery = ("SELECT * FROM sptechnologies  WHERE cecID = \'"+str(cecID)+"\';")
        selectlifeCycleQuery = ("SELECT * FROM migrationLifeCycle  WHERE cecID = \'"+str(cecID)+"\';")
        selectscriptingQuery = ("SELECT * FROM scripting  WHERE cecID = \'"+str(cecID)+"\';")
        selectccieQuery = ("SELECT * FROM activeccie  WHERE cecID = \'"+str(cecID)+"\';")
        selectcompplatformQuery = ("SELECT * FROM competitiveplatforms  WHERE cecID = \'"+str(cecID)+"\';")
        selectSoftQuery = ("SELECT * FROM softskills  WHERE cecID = \'"+str(cecID)+"\';")

    
        resourceDict = selectDB(selectResourceQuery)
        certDict = selectDB(selectCertQuery)
        entDict = selectDB(selectentQuery)
        spDict = selectDB(selectSpQuery)
        migLifeCycleDict = selectDB(selectlifeCycleQuery)
        scriptingDict = selectDB(selectscriptingQuery)
        ccieDict = selectDB(selectccieQuery)
        compPlatformDict = selectDB(selectcompplatformQuery)
        softSkillsDict = selectDB(selectSoftQuery)


        labelEnt = ['Routing','Switching','ItilFoundation','OperationsTools','Design','DnaSDA','Multicast','DMVPN','Wireless',
    'Qos','virtualization','Security','Nexus Hardware','VX Lan','SD WAN','DNAC','Ipv6','Meraki','SASE','SDN','MultiCloud']

        labelSp = ['RoutingSwitching','MplsL2L3Vpn','MplsTE','MulticastMultiCastVpn','XrHardwarePlatformsOs','AccessPlatforms',
            'SegmentRouting','SegmentRoutingSdnBgpPcep','SegmentRoutingTE','EvpnPbBEvpn','UMMT','NSO','BNG','WAN Automation',
            'NFV','LISP','Spit Fire','SRV6','RON','TDM to IP','EPNM','CN BGN']

        labellifecycle = ['Pre Sales ActivityRFP','Pre Sales ActivityTools','Cure Hw/Sw Report','Nmp MasterMop','Hld Document',
            'Lld Document','Crd Questionnair','Customer WorkShop','Nip Document Executed','Nrfu Document','Config Conversion',
            'Mop Document','Migration Execution Support','Lab Testing Poc','Escalation Support','Ic Development']

        labelScripting = ['Python','Vba Scripting','REST Api','Json Xml','Netconf/RestconfYang','Others Scripting Skills']

        labelCcie = ['Routing And Switching','Service Provider','Data Center','Security','Wireless','Collaboration','Ccie Number',
                'Ccie Aspirations']
        labelCompPlat = ['ALU','Juniper','Huawei','Ericsson','HP','IBM']

        labelSoft = ['Capability Devlopment Pgm', 'CDP Foundation','CDP V2','Any Other Softskills','Modified By']



        certList = list()
        rList = list()
        entList = list()
        spList = list()
        miglife = list()
        scriptList = list()
        ccieList = list()
        compPlatformList = list()
        softskillsList = list()

        for k,v in compPlatformDict.items():
            compPlatformList = v

        for k,v in softSkillsDict.items():
            softskillsList = v

        for k,v in resourceDict.items():
            rList = v

        for k,v in entDict.items():
            entList = v

        for k,v in spDict.items():
            spList = v

        for k,v in migLifeCycleDict.items():
            miglife = v

        for k,v in scriptingDict.items():
            for i in range(0,len(v)):
                if i == len(v)-1:
                    scriptList.append(v[i])
                if v[i] == 1:
                    scriptList.append("True")
                else:
                    scriptList.append("False")

        for k,v in ccieDict.items():
            ccieList = v

        for k,v in certDict.items():
            if v[0] == '1':
                certList.append('True')
            else:
                certList.append('Falue')
            certList.append(v[1])
            if v[2] == '1':
                certList.append('True')
            else:
                certList.append('False')
            certList.append(v[3])
    
        entInfo = zip(labelEnt,entList)
        spInfo = zip(labelSp,spList)
        migInfo = zip(labellifecycle,miglife)
        scriptInfo = zip(labelScripting,scriptList)
        ccieInfo = zip(labelCcie,ccieList)
        compPlatInfo = zip(labelCompPlat,compPlatformList)
        softInfo = zip(labelSoft,softskillsList)


        return render(request,'resourceDetails.html',{'devnet':certList[0],'devnetDate':certList[1],'specia':certList[2],'speciaDate':certList[3],'manager':rList[4],'project':rList[2],
        'entsp':rList[3],'incli':rList[5],'name':rList[0],'entInfo':entInfo,'spInfo':spInfo,'migInfo':migInfo,'scriptInfo':scriptInfo
        ,'ccieInfo':ccieInfo,'compPlatInfo':compPlatInfo,'softInfo':softInfo})

    except:
        return HttpResponse("OOPS! Seems no data related to this Employee was uploaded.")
        
    
def ResourcesInfo1(request):

    selectBlue = ("SELECT count(ciscoOrPartner) FROM availabilityTracker  WHERE ciscoOrPartner = 'Cisco';")
    selectRed = ("SELECT count(ciscoOrPartner) FROM availabilityTracker  WHERE ciscoOrPartner <> 'Cisco';")
    selCcieRoute = ("SELECT count(routingAndSwitching) FROM activeccie  WHERE routingAndSwitching =\'"+str(1)+"\';")
    selCcieSp = ("SELECT count(serviceProvider) FROM activeccie  WHERE serviceProvider =\'"+str(1)+"\';")
    selCcieDc = ("SELECT count(dataCenter) FROM activeccie  WHERE dataCenter =\'"+str(1)+"\';")
    selCcieSec = ("SELECT count(security) FROM activeccie  WHERE security =\'"+str(1)+"\';")
    selCcieWireless = ("SELECT count(wireless) FROM activeccie  WHERE wireless =\'"+str(1)+"\';")
    selCcieCollab = ("SELECT count(collaboration) FROM activeccie  WHERE collaboration =\'"+str(1)+"\';")
    selDevnet = ("SELECT count(devnet) FROM othercerts  WHERE devnet =\'"+str(1)+"\';")
    selSDWan = ("SELECT count(specialistEntSdWan) FROM othercerts  WHERE specialistEntSdWan =\'"+str(1)+"\';")
    selPython = ("SELECT count(python) FROM scripting  WHERE python =\'"+str(1)+"\';")
    selVba = ("SELECT count(vbaScripting) FROM scripting  WHERE vbaScripting =\'"+str(1)+"\';")
    selRestApi = ("SELECT count(restApi) FROM scripting  WHERE restApi =\'"+str(1)+"\';")
    selJsonXml = ("SELECT count(jsonXml) FROM scripting  WHERE jsonXml =\'"+str(1)+"\';")
    selNetConf = ("SELECT count(netconfRestconfYang) FROM scripting  WHERE netconfRestconfYang =\'"+str(1)+"\';")
    selectManagerQuery = ("SELECT reportingManager, nce FROM availabilityTracker  WHERE CHAR_LENGTH(nce)<=2;")


    blueCount = selectDB(selectBlue)
    redCount = selectDB(selectRed)
    ccieRouteCount = list(selectDB(selCcieRoute).keys())[0]
    ccieSpCount = list(selectDB(selCcieSp).keys())[0]
    ccieDcCount = list(selectDB(selCcieDc).keys())[0]
    ccieSecCount = list(selectDB(selCcieSec).keys())[0]
    ccieWirelessCount = list(selectDB(selCcieWireless).keys())[0]
    ccieCollabCount = list(selectDB(selCcieCollab).keys())[0]
    devnetCount = list(selectDB(selDevnet).keys())[0]
    sdWanbCount = list(selectDB(selSDWan).keys())[0]
    pythonCount = list(selectDB(selPython).keys())[0]
    vbaCount = list(selectDB(selVba).keys())[0]
    restApiCount = list(selectDB(selRestApi).keys())[0]
    jsonXmlCount = list(selectDB(selJsonXml).keys())[0]
    netConfCount = list(selectDB(selNetConf).keys())[0]
    managerRes = selectDB(selectManagerQuery)

    ccieLabels = ['Routing and Switching','Service Provider','Data Center','Security','Wireless','Collaboration']
    otherCertsLab = ['DevNet', 'Specialist SDWAN','Python','VBA Scripting','REST Api','JSON/XML','Netconf Restconf Yang']

    count = [list(redCount.keys())[0],list(blueCount.keys())[0]]
    ccieCounts = [ccieRouteCount,ccieSpCount,ccieDcCount,ccieSecCount,ccieWirelessCount,ccieCollabCount]
    otherCertCoun = [devnetCount,sdWanbCount,pythonCount,vbaCount,restApiCount,jsonXmlCount,netConfCount]

    managersList = list()
    empCount = list()

    for k,v in managerRes.items():
        managersList.append(k)
        empCount.append(v[0])
    
    count = json.dumps(count)
    ccieLabels = json.dumps(ccieLabels)
    ccieCounts = json.dumps(ccieCounts)
    otherCertsLab = json.dumps(otherCertsLab)
    otherCertCoun = json.dumps(otherCertCoun)
    managersList = json.dumps(managersList)
    empCount = json.dumps(empCount)

    return render(request, 'resources.html',{'count':count,'ccieLabels':ccieLabels,'ccieCounts':ccieCounts,'otherCertsLab':otherCertsLab,
        'otherCertCoun':otherCertCoun,'managersList':managersList,'empCount':empCount})
        
def showTelemetryAvailabilityTracker(request):
    selectManagerQuery = ("SELECT reportingManager, nce FROM availabilityTracker  WHERE CHAR_LENGTH(nce)<=2;")
    selectManagers = selectDB(selectManagerQuery)

    managers = list()
    reporteesCount = list()

    for k,v in selectManagers.items():
        managers.append(k)
        reporteesCount.append(v[0])

    bwDetails = list()
    for manager in managers:
        selectBandwidth =  ("SELECT nce,reportingManager,availabilityMonth1,availabilityMonth2,availabilityMonth3,availabilityMonth4,availabilityMonth5 FROM availabilityTracker  WHERE CHAR_LENGTH(nce)>2 and reportingManager = \'"+str(manager)+"\';")
        m1=0
        m2=0
        m3=0
        m4=0
        m5=0
        summedList = list()
        count = 0
        temp = selectDB(selectBandwidth)
        for k,v in temp.items():
            if v[1] is None:
                m1 += 0
            else:
                m1 += round(float(v[1]) *100,2)
            if v[2] is None:
                m2 += 0
            else:
                m2 += round(float(v[2]) *100,2)
            if v[3] is None:
                m3 += 0
            else:
                m3 += round(float(v[3]) *100,2)
            if v[4] is None:
                m4 += 0
            else:
                m4 += round(float(v[4]) *100,2)
            if v[5] is None:
                m5 += 0
            else:
                m5 += round(float(v[5]) *100,2)
            count += 1

        tempList = [m1/count,m2/count,m3/count,m4/count,m5/count]
        bwDetails.append(tempList)
    mon1 = list()
    mon2 = list()
    mon3 = list()
    mon4 = list()
    mon5 = list()
    for lis in bwDetails:
        mon1.append(round(lis[0],0))
        mon2.append(round(lis[1],0))
        mon3.append(round(lis[2],0))
        mon4.append(round(lis[3],0))
        mon5.append(round(lis[4],0))

    months = ['Jan','Feb','Mar','Apr','May','June','July','Aug','Sep','Oct','Nov','Dec']
    currentMonth = datetime.now().month
    monthsTracker = list()
    months1AvailabilityTracker = list()
    months2AvailabilityTracker = list()
    months3AvailabilityTracker = list()
    months4AvailabilityTracker = list()
    months5AvailabilityTracker = list()

    if currentMonth <= 8:
        for i in range(currentMonth-1,currentMonth+4):
            monthsTracker.append(months[i])
    else:
        for i in range(currentMonth-1,currentMonth+4):
            monthsTracker.append(months[i%len(months)])

    mon11 = monthsTracker[0]
    mon22 = monthsTracker[1]
    mon33 = monthsTracker[2]
    mon44 = monthsTracker[3]
    mon55 = monthsTracker[4]
    
    managers = json.dumps(managers)
    mon1 = json.dumps(mon1)
    mon2 = json.dumps(mon2)
    mon3 = json.dumps(mon3)
    mon4 = json.dumps(mon4)
    mon5 = json.dumps(mon5)
    mon11 = json.dumps(mon11)
    mon22 = json.dumps(mon22)
    mon33 = json.dumps(mon33)
    mon44 = json.dumps(mon44)
    mon55 = json.dumps(mon55)


    return render(request, 'showTelemetryAvailabilityTracker.html',{'managers':managers,'mon1':mon1,'mon2':mon2,'mon3':mon3,'mon4':mon4,'mon5':mon5,'mon11':mon11,'mon22':mon22,'mon33':mon33,'mon44':mon44,'mon55':mon55,'reporteesCount':reporteesCount})